# coding=utf-8
"""
读取测试用例
1. 定义用例文件数据的格式，存储格式，存储位置
2. 读取用例文件
3. 整理数据
4. 返回数据
"""
from openpyxl import load_workbook


def read_excel(file_path, column_rel, sheet_name="001"):
    """
    读取excel文件
    :param file_path: str: 文件路径,格式必须为xlsx
    :param sheet_name: str: sheet名称
    :return: list: [{}]
    """
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    # 读取第一行列头
    key_dict = {}
    for col in range(1,sheet.max_column+1):
        key_dict[col] = column_rel.get(sheet.cell(1,col).value,sheet.cell(1,col).value)
    # 读取数据行
    ret_data = []
    for row in range(2,sheet.max_row+1):
        row_data = {}
        for col in key_dict:
            row_data[key_dict[col]] = sheet.cell(row,col).value
        ret_data.append(row_data)
    return ret_data


def read_csv(file_path):
    """
    读取csv文件
    :param file_path:
    :return:
    """
    return [{}]

def read_file(file_path, **kwargs):
    """
    读取用例文件
    :param file_path:
    :param sheet_name:
    :return:
    """
    if file_path.endswith(".xlsx"):
        data_list = read_excel(file_path, **kwargs)
    elif file_path.endswith(".csv"):
        data_list = read_csv(file_path)
    else:
        data_list = [{}]
    return data_list

def parse_dict(ori_str):
    """
    解析字符串为字典
    :param ori_str: str: "xxx:1\nyyy:2\nzzz:3"
    :return: dict: {
        "xxx": "1",
        "yyy": "2",
        "zzz": "3",
    }
    """
    tmp = ori_str.split("\n")
    ret = {}
    for h in tmp:
        if not h:
            continue
        try:
            k, v = h.split(":")
        except ValueError:
            continue
        ret[k] = v
    return ret


def reorganize(data_list):
    """
    整理数据
    :param data_list: list: [{}]
    :return: list: [{}]
    """
    for data in data_list:
        # 去掉首尾空格
        for k in data:
            if isinstance(data[k], str):
                data[k] = data[k].strip(" ")
        # 整理请求头 参数
        data["headers"] = parse_dict(data.get("headers", ""))
        data["parameter"] = parse_dict(data.get("parameter", ""))
    return data_list


def read_case(case_file_path, column_rel):
    """
    读测试用例
    :param case_file_path: str: 测试用例文件路径
    :param column_rel: str: 字段对应关系
    :return: list: [
        {
            "case_no": "001",
            "url": "",
            "2": "",
            "xxx": "",
        },
    ]
    """
    excel_data_list = read_file(case_file_path, column_rel=column_rel)
    data_list = reorganize(excel_data_list)
    return data_list


if __name__ == "__main__":
    column_ref = {
        "序号": "case_id",
        "接口模块": "module_name",
        "用例标题": "case_title",
        "请求头": "headers",
        "请求方式": "method",
        "接口地址": "url",
        "参数输入": "parameter",
        "期望返回结果": "expect",
        "备注": "remark"
    }
    print(read_case("./case_dir/case_001.xlsx", column_ref))
    print("="*50)
    print(parse_dict(""))
    print(parse_dict("xxx:1\nyyy:2\nzzz3"))
    print(parse_dict("xxx:1\nyyy:2\nzzz:3\n"))
    print(parse_dict(" "))
